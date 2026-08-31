"""Fenêtre "Rapports" : répartition des fonds (% du budget/montant) avec
deux sélecteurs — Pays, puis Répartir par (Catégorie / Pays / Bailleur /
Code comptable) — un tableau et un graphique en barres associé."""

import tkinter as tk
from tkinter import ttk, filedialog

from .. import database as db
from .dialogs import show_info, show_error

DIM_CATEGORY = "Catégorie"
DIM_COUNTRY = "Pays"
DIM_DONOR = "Bailleur (budget)"
DIM_CHARGE_CODE = "Code comptable"

ALL_DIMENSIONS = [DIM_CATEGORY, DIM_COUNTRY, DIM_DONOR, DIM_CHARGE_CODE]
DIMENSIONS_WITHOUT_COUNTRY = [DIM_CATEGORY, DIM_DONOR, DIM_CHARGE_CODE]

BAR_COLORS = ["#4C78A8", "#F58518", "#54A24B", "#E45756", "#B279A2",
              "#9C755F", "#EECA3B", "#72B7B2", "#FF9DA6", "#BAB0AC"]


def _fmt_money(v):
    try:
        return f"{v:,.0f}".replace(",", " ")
    except (TypeError, ValueError):
        return "0"


class ReportsWindow(tk.Toplevel):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.title("Rapports — Répartition des fonds")
        self.geometry("1180x650")
        self.minsize(820, 480)

        self.tree = None
        self.current_rows = []
        self.current_value_key = "pct_budget"

        self._build_toolbar()
        self._build_content()

        self.country_var.set("Tous les pays")
        self._update_dimension_options()
        self.dimension_var.set(DIM_CATEGORY)
        self.refresh()

    def _build_toolbar(self):
        bar = ttk.Frame(self, padding=8)
        bar.pack(fill="x")

        ttk.Label(bar, text="Pays :").pack(side="left", padx=(0, 4))
        self.country_var = tk.StringVar()
        countries = ["Tous les pays"] + [c["name"] for c in db.list_countries(self.conn)]
        self.country_combo = ttk.Combobox(
            bar, textvariable=self.country_var, values=countries, state="readonly", width=16,
        )
        self.country_combo.pack(side="left", padx=(0, 16))
        self.country_combo.bind("<<ComboboxSelected>>", self._on_country_change)

        ttk.Label(bar, text="Répartir par :").pack(side="left", padx=(0, 4))
        self.dimension_var = tk.StringVar()
        self.dimension_combo = ttk.Combobox(
            bar, textvariable=self.dimension_var, values=ALL_DIMENSIONS, state="readonly", width=20,
        )
        self.dimension_combo.pack(side="left")
        self.dimension_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        ttk.Button(bar, text="Exporter ce rapport (.xlsx)", command=self._export_report).pack(side="right")

    def _on_country_change(self, event=None):
        self._update_dimension_options()
        self.refresh()

    def _update_dimension_options(self):
        if self.country_var.get() == "Tous les pays":
            options = ALL_DIMENSIONS
        else:
            options = DIMENSIONS_WITHOUT_COUNTRY
        self.dimension_combo["values"] = options
        if self.dimension_var.get() not in options:
            self.dimension_var.set(options[0])

    def _selected_country_id(self):
        name = self.country_var.get()
        if not name or name == "Tous les pays":
            return None
        row = self.conn.execute("SELECT id FROM countries WHERE name = ?", (name,)).fetchone()
        return row["id"] if row else None

    def _build_content(self):
        split = ttk.PanedWindow(self, orient="horizontal")
        split.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.table_frame = ttk.Frame(split)
        self.chart_frame = ttk.LabelFrame(split, text="Graphique")
        split.add(self.table_frame, weight=2)
        split.add(self.chart_frame, weight=3)

        self.canvas = tk.Canvas(self.chart_frame, bg="white", highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind("<Configure>", lambda e: self._draw_chart())

    def _make_tree(self, columns, headers, widths):
        if self.tree is not None:
            self.tree.destroy()
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show="headings")
        for c, h, w in zip(columns, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        vsb = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.tag_configure("total", font=("Segoe UI", 9, "bold"))

    def refresh(self):
        dimension = self.dimension_var.get()
        country_id = self._selected_country_id()

        if dimension == DIM_CATEGORY:
            rows = db.breakdown_by_category(self.conn, country_id)
            self._fill_activity_style_table("Catégorie", rows)
            self.current_value_key = "pct_budget"
        elif dimension == DIM_COUNTRY:
            rows = db.breakdown_by_country(self.conn)
            self._fill_activity_style_table("Pays", rows)
            self.current_value_key = "pct_budget"
        elif dimension == DIM_DONOR:
            rows = db.breakdown_by_donor(self.conn, country_id)
            self._fill_donor_style_table(rows)
            self.current_value_key = "pct_budget"
        elif dimension == DIM_CHARGE_CODE:
            rows = db.breakdown_by_charge_code_pct(self.conn, country_id)
            self._fill_charge_code_style_table(rows)
            self.current_value_key = "pct_montant"
        else:
            rows = []

        self.current_rows = rows
        self._draw_chart()

    def _fill_activity_style_table(self, label_header, rows):
        columns = ("label", "budget", "pct_budget", "n_activities", "pct_activities")
        headers = [label_header, "Budget", "% du budget", "Activités", "% des activités"]
        widths = [160, 130, 100, 90, 110]
        self._make_tree(columns, headers, widths)
        for r in rows:
            tags = ("total",) if r.get("is_total") else ()
            self.tree.insert("", "end", values=(
                r["label"], _fmt_money(r["budget"]), f"{r['pct_budget']:.1f}%",
                r["n_activities"], f"{r['pct_activities']:.1f}%",
            ), tags=tags)

    def _fill_donor_style_table(self, rows):
        columns = ("label", "budget", "pct_budget")
        headers = ["Bailleur", "Budget", "% du budget"]
        widths = [160, 150, 120]
        self._make_tree(columns, headers, widths)
        for r in rows:
            tags = ("total",) if r.get("is_total") else ()
            self.tree.insert("", "end", values=(
                r["label"], _fmt_money(r["budget"]), f"{r['pct_budget']:.1f}%",
            ), tags=tags)

    def _fill_charge_code_style_table(self, rows):
        columns = ("label", "montant", "pct_montant")
        headers = ["Code comptable", "Montant achats", "% du montant"]
        widths = [180, 150, 120]
        self._make_tree(columns, headers, widths)
        for r in rows:
            tags = ("total",) if r.get("is_total") else ()
            self.tree.insert("", "end", values=(
                r["label"], _fmt_money(r["montant"]), f"{r['pct_montant']:.1f}%",
            ), tags=tags)

    def _draw_chart(self):
        canvas = self.canvas
        canvas.delete("all")

        data = [r for r in self.current_rows if not r.get("is_total")]
        if not data:
            canvas.create_text(20, 20, anchor="nw", text="Aucune donnée à afficher.")
            return

        width = max(canvas.winfo_width(), 300)
        height = max(canvas.winfo_height(), 260)

        margin_left = 46
        margin_bottom = 90
        margin_top = 26
        chart_h = max(height - margin_top - margin_bottom, 40)
        chart_w = max(width - margin_left - 20, 60)

        n = len(data)
        bar_w = min(60, chart_w / max(n, 1) * 0.55)
        gap = (chart_w - bar_w * n) / (n + 1) if n else 0

        for pct in (0, 25, 50, 75, 100):
            y = margin_top + chart_h - (pct / 100) * chart_h
            canvas.create_line(margin_left, y, width - 10, y, fill="#e8e8e8")
            canvas.create_text(margin_left - 6, y, anchor="e", text=f"{pct}%", font=("Segoe UI", 7), fill="#888")

        x = margin_left + gap
        for i, item in enumerate(data):
            val = item.get(self.current_value_key, 0) or 0
            bar_h = (val / 100) * chart_h
            y1 = margin_top + chart_h
            y0 = y1 - bar_h
            color = BAR_COLORS[i % len(BAR_COLORS)]
            cap_h = min(10, bar_w / 2, max(bar_h, 1))

            canvas.create_rectangle(x, y0 + cap_h / 2, x + bar_w, y1, fill=color, outline="")
            canvas.create_oval(x, y0, x + bar_w, y0 + cap_h, fill=color, outline="")
            canvas.create_oval(x, y1 - cap_h / 2, x + bar_w, y1 + cap_h / 2, fill=color, outline="")

            canvas.create_text(x + bar_w / 2, y0 - 10, text=f"{val:.1f}%", font=("Segoe UI", 8, "bold"))

            label = str(item["label"])
            if len(label) > 18:
                label = label[:17] + "…"
            canvas.create_text(x + bar_w / 2, y1 + 14, text=label, font=("Segoe UI", 7), angle=35, anchor="e")
            x += bar_w + gap

        canvas.create_line(margin_left, margin_top, margin_left, margin_top + chart_h, fill="#bbb")
        canvas.create_line(margin_left, margin_top + chart_h, width - 10, margin_top + chart_h, fill="#bbb")

    def _export_report(self):
        if self.tree is None or not self.tree.get_children():
            show_info(self, "Rapport vide", "Aucune donnée à exporter pour ce rapport.")
            return

        path = filedialog.asksaveasfilename(
            title="Exporter le rapport",
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile=f"rapport_{self.dimension_var.get().lower().replace(' ', '_').replace('(', '').replace(')', '')}.xlsx",
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapport"

            headers = [self.tree.heading(c)["text"] for c in self.tree["columns"]]
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="1F4E78")

            for r_idx, item in enumerate(self.tree.get_children(), start=2):
                values = self.tree.item(item, "values")
                for col, v in enumerate(values, start=1):
                    ws.cell(row=r_idx, column=col, value=v)

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

            wb.save(path)
        except Exception as exc:  # noqa: BLE001
            show_error(self, "Erreur d'export", str(exc))
            return

        show_info(self, "Export terminé", f"Le rapport a été enregistré :\n{path}")
