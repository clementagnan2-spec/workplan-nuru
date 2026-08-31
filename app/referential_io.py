"""
Import et modèle Excel pour les référentiels (Codes d'activité, Catégories,
Budget/bailleurs, Codes de charge). Format à 2 colonnes : Code (ou Nom) /
Libellé. Réimportable sans risque : une valeur déjà présente est mise à
jour (libellé remplacé), pas dupliquée.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import database as db

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_referential_template(path: str, col1_label: str, col2_label: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Référentiel"

    for idx, h in enumerate((col1_label, col2_label), start=1):
        c = ws.cell(row=1, column=idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=1, value="Exemple")
    ws.cell(row=2, column=2, value="Libellé de l'exemple — remplacez ou supprimez cette ligne")

    for i in range(1, 3):
        ws.column_dimensions[get_column_letter(i)].width = 30

    wb.save(path)


def import_referential(conn, table: str, path: str) -> int:
    """Importe un fichier à 2 colonnes (Code/Nom, Libellé) dans le
    référentiel `table`. Une valeur déjà présente voit son libellé mis à
    jour plutôt que d'être dupliquée. Retourne le nombre de lignes traitées."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    key = db.REFERENTIAL_TABLES[table]
    existing = {row[key]: row["id"] for row in db.list_referential(conn, table)}

    n = 0
    for r in range(2, ws.max_row + 1):
        raw_value = ws.cell(row=r, column=1).value
        raw_label = ws.cell(row=r, column=2).value
        if raw_value is None or str(raw_value).strip() == "":
            continue
        value = str(raw_value).strip()
        label = str(raw_label).strip() if raw_label is not None else None

        if value in existing:
            db.update_referential(conn, table, existing[value], value, label)
        else:
            new_id = db.add_referential(conn, table, value, label)
            existing[value] = new_id
        n += 1

    return n
