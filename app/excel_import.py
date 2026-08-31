"""
Import d'un classeur Excel de type "WORKPLAN_MULTIPAYS" vers la base de
données SQLite de l'application.

L'avancement (colonne "AVANCEMENT" du fichier source, si présente) n'est
PAS importé : il est toujours recalculé automatiquement par
database.add_activity() comme (coût total / budget total).
"""

import re
import datetime
import openpyxl

from . import database as db

PROCUREMENT_COLUMN_ORDER = db.PROCUREMENT_FIELDS


def _to_iso_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.date().isoformat() if isinstance(value, datetime.datetime) else value.isoformat()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s or None


def _to_float(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except ValueError:
        return 0.0


def _to_text(value):
    if value is None:
        return None
    return str(value).strip() or None


PHASE_RE = re.compile(r"^\s*Phase\s+\d+", re.IGNORECASE)
END_MARK_RE = re.compile(r"marque la fin du planning", re.IGNORECASE)


def import_workbook(conn, path: str, progress_callback=None) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    summary = {}

    for sheet_name in wb.sheetnames:
        low = sheet_name.strip().lower()
        if low.startswith("workplan") or low.startswith("workpla"):
            country = _guess_country(sheet_name)
            if not country:
                continue
            if progress_callback:
                progress_callback(f"Import planning {country}...")
            n_act = _import_workplan_sheet(conn, wb[sheet_name], country)
            summary.setdefault(country, {})["activities"] = n_act
        elif low.startswith("procurement") or low.startswith("procureme"):
            country = _guess_country(sheet_name)
            if not country:
                continue
            if progress_callback:
                progress_callback(f"Import achats {country}...")
            n_proc = _import_procurement_sheet(conn, wb[sheet_name], country)
            summary.setdefault(country, {})["procurements"] = n_proc

    # Les coûts des activités ne sont jamais lus depuis le fichier : ils sont
    # recalculés à partir des achats au statut "Livré" pour tous les pays,
    # pour rester cohérents même si l'import ne couvre qu'une partie du classeur.
    if progress_callback:
        progress_callback("Ventilation des coûts depuis les achats livrés...")
    for country in db.list_countries(conn):
        db.recompute_costs_from_procurements(conn, country["id"])

    return summary


def _guess_country(sheet_name: str):
    known = ["TOGO", "BENIN", "BÉNIN", "NIGER", "GHANA", "GHNANA"]
    upper = sheet_name.upper()
    for c in known:
        if c in upper:
            return "GHANA" if c == "GHNANA" else ("BENIN" if c == "BÉNIN" else c)
    return None


def _find_header_row(ws, marker_text: str, col: int = 1, max_scan: int = 15):
    for r in range(1, max_scan + 1):
        val = ws.cell(row=r, column=col).value
        if val and marker_text.lower() in str(val).lower():
            return r
    return None


def _import_workplan_sheet(conn, ws, country_name: str) -> int:
    country_id = db.get_or_create_country(conn, country_name)
    header_row = _find_header_row(ws, "RETARD")
    if header_row is None:
        return 0

    current_phase_id = None
    current_phase_pos = 0
    n_imported = 0

    for r in range(header_row + 1, ws.max_row + 1):
        col_a = ws.cell(row=r, column=1).value
        col_b = ws.cell(row=r, column=2).value
        col_c = ws.cell(row=r, column=3).value

        if col_a and END_MARK_RE.search(str(col_a)):
            break

        if col_b and PHASE_RE.match(str(col_b)):
            current_phase_pos += 1
            phase_label = f"{str(col_b).strip()} - {str(col_c).strip()}" if col_c else str(col_b).strip()
            current_phase_id = db.get_or_create_phase(conn, country_id, phase_label, current_phase_pos)
            continue

        if not col_b and not col_c:
            continue
        if col_a and "insérez" in str(col_a).lower():
            continue

        code = _to_text(col_b)
        task = _to_text(col_c) or code or "(sans nom)"

        debut = ws.cell(row=r, column=5).value
        fin = ws.cell(row=r, column=6).value
        nb_pieces = ws.cell(row=r, column=7).value

        cost_p = ws.cell(row=r, column=8).value
        cost_i = ws.cell(row=r, column=9).value
        cost_a = ws.cell(row=r, column=10).value
        cost_c = ws.cell(row=r, column=11).value

        budget_ni = ws.cell(row=r, column=17).value
        budget_tifr = ws.cell(row=r, column=18).value
        budget_ftit = ws.cell(row=r, column=19).value

        cat_values = {"P": _to_float(cost_p), "I": _to_float(cost_i),
                      "A": _to_float(cost_a), "C": _to_float(cost_c)}
        category = max(cat_values, key=cat_values.get) if any(cat_values.values()) else None

        data = {
            "phase_id": current_phase_id,
            "code": code,
            "task": task,
            "assigned_to": None,
            # "progress" n'est pas lu depuis le fichier : calculé automatiquement
            # "cost_*" non plus : recalculé après import depuis les achats "Livré"
            # (voir database.recompute_costs_from_procurements, appelé en fin d'import)
            "start_date": _to_iso_date(debut),
            "end_date": _to_iso_date(fin),
            "nb_pieces": _to_float(nb_pieces),
            "category": category,
            "budget_ni_hct": _to_float(budget_ni),
            "budget_tifr_usaid": _to_float(budget_tifr),
            "budget_ftit": _to_float(budget_ftit),
            "comment": None,
        }
        db.add_activity(conn, country_id, data)
        n_imported += 1

    return n_imported


def _import_procurement_sheet(conn, ws, country_name: str) -> int:
    country_id = db.get_or_create_country(conn, country_name)
    header_row = _find_header_row(ws, "Lien Dossier WORK PLAN")
    if header_row is None:
        return 0

    n_imported = 0
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=r, column=c + 1).value for c in range(len(PROCUREMENT_COLUMN_ORDER))]
        if all(v is None for v in values):
            continue

        data = {}
        for field, raw in zip(PROCUREMENT_COLUMN_ORDER, values):
            if field == "montant":
                data[field] = _to_float(raw)
            elif field in ("date_bc", "date_fournisseur", "date_livraison_prevue",
                           "date_reception_facture", "date_paiement"):
                data[field] = _to_iso_date(raw)
            else:
                data[field] = _to_text(raw)

        if not data.get("designation") and not data.get("n_bc") and not data.get("dossier_workplan"):
            continue

        db.add_procurement(conn, country_id, data)
        n_imported += 1

    return n_imported
