"""Export de la base de données de l'application vers un classeur Excel."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import database as db

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _write_header(ws, row, headers):
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, n_cols, width=16):
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def export_workbook(conn, path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    consolidation = wb.create_sheet("Consolidation")
    consolidation["A1"] = "Consolidation multi-pays — Nuru Workplan Manager"
    consolidation["A1"].font = TITLE_FONT
    headers = ["Pays", "Coût NI/HCT", "Coût TIFR-USAID", "Coût FTIT", "Coût total",
               "Budget NI/HCT", "Budget TIFR-USAID", "Budget FTIT", "Budget total", "Solde"]
    _write_header(consolidation, 3, headers)

    row = 4
    countries = db.list_countries(conn)
    for country in countries:
        totals = db.country_totals(conn, country["id"])
        consolidation.cell(row=row, column=1, value=country["name"])
        consolidation.cell(row=row, column=2, value=totals["cost_ni_hct"])
        consolidation.cell(row=row, column=3, value=totals["cost_tifr_usaid"])
        consolidation.cell(row=row, column=4, value=totals["cost_ftit"])
        consolidation.cell(row=row, column=5, value=totals["cost_total"])
        consolidation.cell(row=row, column=6, value=totals["budget_ni_hct"])
        consolidation.cell(row=row, column=7, value=totals["budget_tifr_usaid"])
        consolidation.cell(row=row, column=8, value=totals["budget_ftit"])
        consolidation.cell(row=row, column=9, value=totals["budget_total"])
        consolidation.cell(row=row, column=10, value=totals["solde"])
        row += 1

    _autosize(consolidation, len(headers))

    for country in countries:
        _export_country_planning(wb, conn, country)
        _export_country_procurement(wb, conn, country)

    wb.save(path)


def _export_country_planning(wb, conn, country):
    ws = wb.create_sheet(f"Planning {country['name']}"[:31])
    headers = ["Phase", "Code", "Tâche", "Assigné à", "Avancement (%)",
               "Début", "Fin", "Nb pièces", "Catégorie",
               "Coût NI/HCT", "Coût TIFR-USAID", "Coût FTIT", "Coût total",
               "Budget NI/HCT", "Budget TIFR-USAID", "Budget FTIT", "Budget total",
               "Solde", "Non livré", "Solde global", "Commentaire"]
    _write_header(ws, 1, headers)

    row = 2
    for a in db.list_activities(conn, country["id"]):
        cost_total = a["cost_ni_hct"] + a["cost_tifr_usaid"] + a["cost_ftit"]
        budget_total = a["budget_ni_hct"] + a["budget_tifr_usaid"] + a["budget_ftit"]
        solde = budget_total - cost_total
        non_livre = a["non_delivered"] or 0
        values = [
            a["phase_name"], a["code"], a["task"], a["assigned_to"],
            round((a["progress"] or 0) * 100, 1), a["start_date"], a["end_date"],
            a["nb_pieces"], a["category"],
            a["cost_ni_hct"], a["cost_tifr_usaid"], a["cost_ftit"], cost_total,
            a["budget_ni_hct"], a["budget_tifr_usaid"], a["budget_ftit"], budget_total,
            solde, non_livre, solde - non_livre, a["comment"],
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)
        row += 1

    _autosize(ws, len(headers), width=15)


def _export_country_procurement(wb, conn, country):
    ws = wb.create_sheet(f"Achats {country['name']}"[:31])
    headers = ["Dossier Workplan", "N° PR", "N° RFQ", "N° BC", "Date BC", "N° Proforma",
               "Demandeur", "Désignation", "Fournisseur", "Date fournisseur", "Montant",
               "Catégorie", "Lieu livraison", "Date livraison prévue", "Statut BC",
               "Type procurement", "Projet", "Charge code", "Code", "Bon de livraison",
               "Facture définitive", "Date réception facture", "RIB", "Banque",
               "Mode paiement", "Date paiement", "Validation", "Commentaires"]
    _write_header(ws, 1, headers)

    row = 2
    for p in db.list_procurements(conn, country["id"]):
        values = [
            p["dossier_workplan"], p["n_pr"], p["n_rfq"], p["n_bc"], p["date_bc"],
            p["n_proforma"], p["demandeur"], p["designation"], p["fournisseur"],
            p["date_fournisseur"], p["montant"], p["categorie"], p["lieu_livraison"],
            p["date_livraison_prevue"], p["statut_bc"], p["type_procurement"],
            p["project"], p["charge_code"], p["code"], p["bon_livraison"],
            p["facture_definitive"], p["date_reception_facture"], p["rib"],
            p["banque"], p["mode_paiement"], p["date_paiement"], p["validation"],
            p["commentaires"],
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)
        row += 1

    _autosize(ws, len(headers), width=15)
