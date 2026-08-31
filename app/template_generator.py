"""
Génère un classeur Excel VIERGE (modèle) à la structure attendue par
excel_import.py : une feuille "WORKPLAN <PAYS>" (planning) et une feuille
"procurement <PAYS>" (achats) pour chacun des 4 pays. L'utilisateur le
télécharge, le remplit, puis l'importe via Fichier > Importer.

La colonne AVANCEMENT est conservée dans l'en-tête du modèle pour rester
fidèle au fichier source historique, mais elle n'est jamais lue à
l'import : l'avancement est toujours recalculé automatiquement par
l'application (coût total / budget total).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import database as db
from . import sample_data as sd

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(italic=True, color="888888")

WORKPLAN_HEADERS = [
    "RETARD", "TÂCHE", "ATTRIBUÉE À", "AVANCEMENT (auto - ne pas remplir)", "DÉBUT", "FIN",
    "Nb pieces", "ACTIVITE", "Sensibilisation", "Admin", "Collecte", "TOTAL",
    "Coût / dépenses NI/HCT", "TIFR-USAID", "FTIT", "TOTAL",
    "BUDGET BAILLEURS NI/HCT", "TIFR-USAID", "FTIT", "TOTAL", "SOLDE", "JOURS",
]

PROCUREMENT_HEADERS = [
    "Lien Dossier WORK PLAN", "N° PR", "N° RFQ", "N° Dossier BC", "Date", "N° Proforma",
    "Demandeur", "Designation", "Fournisseur/Vendeur", "DATE", "Montant",
    "Programme (P/I/A/C)", "Lieu de livraison", "Date de livraison prevue",
    "Statut du BC", "Type de procurement", "PROJECT (Bailleur : NI/HCT, TIFR-USAID ou FTIT)",
    "CHARGE CODE", "CODE",
    "Bon de Livraisons", "Facture definitive", "Date de reception facture definitive",
    "RIB", "Banque du fournisseur", "Mode de Paiement", "Date de paiement",
    "Validation/Authorization", "Commentaires",
]


def _write_header(ws, row, headers):
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, n, width=15):
    for i in range(1, n + 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_template(path: str):
    """Crée un classeur vierge avec une ligne d'exemple (à remplacer) par
    pays, pour montrer le format attendu sans imposer de vraies données."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for country in db.DEFAULT_COUNTRIES:
        ws = wb.create_sheet(f"WORKPLAN {country}")
        _write_header(ws, 1, WORKPLAN_HEADERS)

        ws.cell(row=2, column=2, value="Phase 1")
        ws.cell(row=2, column=3, value="Nom de la phase")

        ws.cell(row=3, column=2, value="A1XX")
        ws.cell(row=3, column=3, value="Exemple d'activité à remplacer")
        ws.cell(row=3, column=5, value="2026-01-01")
        ws.cell(row=3, column=6, value="2026-01-15")
        ws.cell(row=3, column=7, value=1)
        ws.cell(row=3, column=8, value=0)      # coût Programme
        ws.cell(row=3, column=9, value=0)      # coût Sensibilisation
        ws.cell(row=3, column=10, value=0)     # coût Admin
        ws.cell(row=3, column=11, value=0)     # coût Collecte
        ws.cell(row=3, column=13, value=0)     # coût NI/HCT
        ws.cell(row=3, column=14, value=0)     # coût TIFR-USAID
        ws.cell(row=3, column=15, value=0)     # coût FTIT
        ws.cell(row=3, column=17, value=0)     # budget NI/HCT
        ws.cell(row=3, column=18, value=0)     # budget TIFR-USAID
        ws.cell(row=3, column=19, value=0)     # budget FTIT

        note = ws.cell(row=5, column=1, value=(
            "Ajoutez vos lignes « Phase X » puis vos activités en dessous. "
            "Les colonnes AVANCEMENT et COÛT ne sont pas utilisées : l'avancement "
            "est calculé automatiquement (coût / budget) et le coût est ventilé "
            "automatiquement depuis les achats au statut « Livré » (voir la feuille "
            "procurement, colonne PROJECT = bailleur)."
        ))
        note.font = NOTE_FONT

        ws.cell(row=7, column=1, value="Cette ligne marque la fin du planning de projet")
        _autosize(ws, len(WORKPLAN_HEADERS))

    for country in db.DEFAULT_COUNTRIES:
        ws = wb.create_sheet(f"procurement {country}")
        _write_header(ws, 1, PROCUREMENT_HEADERS)

        ws.cell(row=2, column=1, value="A1XX")
        ws.cell(row=2, column=8, value="Exemple - désignation de l'achat à remplacer")
        ws.cell(row=2, column=11, value=0)
        ws.cell(row=2, column=15, value="En cours")  # Statut du BC (passer à "Livré" pour ventiler le coût)
        ws.cell(row=2, column=17, value="NI/HCT")    # PROJECT = bailleur (NI/HCT, TIFR-USAID ou FTIT)

        note = ws.cell(row=4, column=len(PROCUREMENT_HEADERS) + 2,
                        value="Insérez vos achats à partir de la ligne 2 (remplacez la ligne d'exemple).")
        note.font = NOTE_FONT
        _autosize(ws, len(PROCUREMENT_HEADERS))

    wb.save(path)
